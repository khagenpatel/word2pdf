import openpyxl
import re
import tkinter as tk
from tkinter import filedialog
from datetime import datetime


# Function to remove line breaks and Batch information
def clean_batch_info(text):
    # Remove line breaks
    text = text.replace('\n', ' ')

    # Remove "Batch" and everything after that
    batch_pattern = r"Batch.*"
    text = re.sub(batch_pattern, '', text)

    return text.strip()


# Create a Tkinter root window
root = tk.Tk()
root.withdraw()  # Hide the root window

# Ask the user to select the input file using a file dialog
input_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

# Load the workbook
workbook = openpyxl.load_workbook(input_file_path)
# Access the specific sheet, e.g., 'Sheet1'
sheet = workbook['Sheet1']

# Initialize the new columns with empty strings
sheet['C1'] = 'Batch Number'
sheet['D1'] = 'Mfg Date'

# Regular expression patterns to search for Batch Number and Mfg Date
batch_pattern = r"Batch No.:(.+)"
mfg_pattern = r"Mfg\. Dt\.(\d{1,2}-\d{4})"

# Loop through each row in the sheet
for row_idx, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    batch_info = clean_batch_info(str(row_values[1]))  # Get the Batch info from the original column B
    sheet.cell(row=row_idx, column=2, value=batch_info)  # Update the cleaned Batch info in column B

    for cell_value in row_values:
        batch_match = re.search(batch_pattern, str(cell_value))
        mfg_match = re.search(mfg_pattern, str(cell_value))

        # Check if Batch Number is found and update the cell in column C
        if batch_match:
            batch_number = batch_match.group(1).strip()
            sheet.cell(row=row_idx, column=3, value=batch_number)

        # Check if Mfg Date is found and update the cell in column D
        if mfg_match:
            mfg_date = mfg_match.group(1).strip()
            mfg_date_obj = datetime.strptime(mfg_date, "%m-%Y")  # Parse date
            formatted_mfg_date = mfg_date_obj.strftime("%b-%Y")  # Format date
            sheet.cell(row=row_idx, column=4, value=formatted_mfg_date)

# Get the current date and time for the output file name
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
output_file_name = f"modified_file_{current_datetime}.xlsx"

# Save the modified workbook
workbook.save(output_file_name)

# Notify the user that the process is complete
print(f"File {output_file_name} has been created with the modified data.")
